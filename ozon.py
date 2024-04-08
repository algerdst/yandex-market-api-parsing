import openpyxl

import requests
import json
import datetime


with open('Импорт ozon API - логин.txt', 'r', encoding='utf-8') as file:
    data = [i for i in file]
client_id = data[0].replace('\n', '')
api_key = data[1].replace('\n', '')
headers = {
    "Client-Id": client_id,
    "Api-Key": api_key
}

url = 'https://api-seller.ozon.ru/v3/posting/fbs/list'
yesterday = datetime.date.today()- datetime.timedelta(days=20)
after_10_days = yesterday + datetime.timedelta(days=20)
payload = json.dumps({
    "dir": "ASC",
    "filter": {

        "order_id": 0,
        "since": f"{yesterday}T11:47:39.878Z",
        "to": f"{after_10_days}T11:47:39.878Z",
    },
    "limit": 1000,


})
resp_data = requests.post(url, headers=headers, data=payload).json()
# with open('file.json', 'w', encoding='utf-8') as file:
#     json.dump(resp_data, file, indent=4, ensure_ascii=False)

today = datetime.date.today()
tomorrow = today + datetime.timedelta(days=1)
date_name=str(tomorrow).split('-')
date_name=f"{date_name[2]}-{date_name[1]}-{date_name[0]}".replace('-','.')

filename=f'Заказы {str(date_name)}.xlsx'

def ozon():
    book = openpyxl.load_workbook(filename)
    sheet=book.active
    row=1

    for i in resp_data['result']['postings']:
        date_time_str=i['shipment_date'][:10]
        date= datetime.datetime.strptime(date_time_str, '%Y-%m-%d')
        date=date.date()
        if date==tomorrow:
            date=str(date).split('-')
            date = f"{date[2]}-{date[1]}-{date[0]}"
            sku=i['products'][0]['offer_id']
            quantity=i['products'][0]['quantity']
            posting_number=i['posting_number']
            sheet.cell(column=6, row=row).value = sku
            sheet.cell(column=7, row=row).value = quantity
            sheet.cell(column=8, row=row).value = str(date).replace('-','.')
            sheet.cell(column=9, row=row).value = 'FBS'
            sheet.cell(column=10, row=row).value = i['delivery_method']['warehouse']
            sheet.cell(column=11, row=row).value = posting_number
            row+=1
            book.save(filename)
    book.close()


