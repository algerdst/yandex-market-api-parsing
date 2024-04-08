import datetime
import json

import openpyxl

import requests
from ozon import ozon, filename
import os
import glob


with open('Импорт YaMarket API - логин.txt', 'r', encoding='utf-8') as file:
    data = [i for i in file]
campaign_id = data[0].replace('\n', '')
token = data[1].replace('\n', '')

date_from=str(datetime.datetime.today().date()).split('-')
date_to=str(datetime.datetime.today().date()+datetime.timedelta(7)).split('-')
date=str(datetime.datetime.today().date()+datetime.timedelta(1)).split('-')
date=f"{date[2]}-{date[1]}-{date[0]}"
date_from=f"{date_from[2]}-{date_from[1]}-{date_from[0]}"
date_to=f"{date_to[2]}-{date_to[1]}-{date_to[0]}"

def delete_xlsx():
    path = os.getcwd()
    for file in glob.glob(os.path.join(path, '*.xlsx')):
        os.remove(file)
delete_xlsx()

url=f'https://api.partner.market.yandex.ru/campaigns/{campaign_id}/orders?status=&substatus=&fromDate=&toDate=&supplierShipmentDateFrom={date_from}&supplierShipmentDateTo={date_to}&updatedAtFrom=&updatedAtTo=&dispatchType=&fake=&hasCis=&onlyWaitingForCancellationApprove=&onlyEstimatedDelivery=&page=&pageSize='

headers={
        "Authorization":f"{token}"
}


response=requests.get(url, headers=headers).json()
pages=int(response['pager']['pagesCount'])

wb = openpyxl.Workbook(filename)
wb.save(filename)
book = openpyxl.load_workbook(filename)
sheet=book.active
row=1
dict_res={}
index=0
for i in range(pages):
    if i==0:
        response=response
    else:
        page=i+1
        url=f"https://api.partner.market.yandex.ru/campaigns/{campaign_id}/orders?status=&substatus=&fromDate=&toDate=&supplierShipmentDateFrom={date_from}&supplierShipmentDateTo={date_to}&updatedAtFrom=&updatedAtTo=&dispatchType=&fake=&hasCis=&onlyWaitingForCancellationApprove=&onlyEstimatedDelivery=&page={page}&pageSize="
        response = requests.get(url, headers=headers).json()
    for i in response['orders']:
        if i['status']=='CANCELLED':
            continue
        shipment_date=i['delivery']['shipments'][0]['shipmentDate']
        if shipment_date!=date:
            continue
        dict_res[index]=i
        index+=1
        item=i['items'][0]
        sku=item['offerId']
        quantity=item['count']
        sheet.cell(column=1, row=row).value = sku
        sheet.cell(column=2, row=row).value = quantity
        sheet.cell(column=3, row=row).value = date.replace('-','.')
        sheet.cell(column=4, row=row).value = 'FBS'
        row += 1
        book.save(filename)

# with open('file.json', 'w', encoding='utf-8') as file:
#     json.dump(dict_res, file, indent=4, ensure_ascii=False)

ozon()
